import os
import datetime
import uuid
import pickle
import faiss
import openai
from flask import Flask, render_template, request, jsonify, session, redirect, url_for
from flask_session import Session
from sentence_transformers import SentenceTransformer
from openpyxl import Workbook, load_workbook
from PyPDF2 import PdfReader
from docx import Document
import time
import threading

# ====== CONFIG & INITIALIZATION ======
# Use environment variable for the API key for security
openai.api_key = os.environ.get("OPENAI_API_KEY")

# Check if the API key is set
if not openai.api_key:
    # This will prevent the app from starting if the key is missing
    raise ValueError("OPENAI_API_KEY environment variable not set.")

# Correctly define the folders based on your description
DATA_FOLDER = "data"    # This folder is for documents to be trained
DATA_FOLDER2 = "data2"  # This folder is for the generated index files and logs

# Corrected file paths to use the new DATA_FOLDER2
INDEX_FILE = os.path.join(DATA_FOLDER2, "faiss_index.index")
MAPPING_FILE = os.path.join(DATA_FOLDER2, "doc_mapping.pkl")
EXCEL_LOG = os.path.join(DATA_FOLDER2, "ChatLogs.xlsx")
TIMESTAMP_FILE = os.path.join(DATA_FOLDER2, "last_training_timestamp.txt")
INDEX_REQUEST_FILE = os.path.join(DATA_FOLDER2, "start_indexing.txt")
SYSTEM_PROMPT_FILE = "system_prompt.txt"

MODEL_NAME = "all-MiniLM-L6-v2"

# Initialize Flask App
app = Flask(__name__)
app.config['SESSION_TYPE'] = 'filesystem'
Session(app)

# --- App-wide data loading ---
index = None
documents_list = []
doc_mapping = {}
embedder = SentenceTransformer(MODEL_NAME)
system_prompt = "You are a helpful assistant."


def load_index_and_docs():
    global index, documents_list, doc_mapping
    print("Loading FAISS index and documents...")
    try:
        if not os.path.exists(INDEX_FILE) or not os.path.exists(MAPPING_FILE):
            raise FileNotFoundError
        index = faiss.read_index(INDEX_FILE)
        with open(MAPPING_FILE, "rb") as f:
            doc_mapping = pickle.load(f)
        documents_list = [item['content'] for item in doc_mapping.values()]
        print("Loaded successfully.")
    except (FileNotFoundError, EOFError):
        print("Error: Index or mapping files not found or corrupt. Please run training to rebuild.")
        index = None
        documents_list = []
        doc_mapping = {}


def load_system_prompt():
    global system_prompt
    try:
        with open(SYSTEM_PROMPT_FILE, "r", encoding="utf-8") as f:
            system_prompt = f.read().strip()
            print(f"System prompt loaded from {SYSTEM_PROMPT_FILE}")
    except FileNotFoundError:
        print(f"Warning: {SYSTEM_PROMPT_FILE} not found. Using default system prompt.")
    except Exception as e:
        print(f"Error loading system prompt: {e}. Using default.")


# --- Document Processing Functions ---
def load_txt(path):
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        return f.read().strip()


def load_pdf(path):
    reader = PdfReader(path)
    texts = [page.extract_text().strip() for page in reader.pages if page.extract_text()]
    return "\n".join(texts)


def load_docx(path):
    doc = Document(path)
    texts = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
    return "\n".join(texts)


def get_file_status():
    # This function now correctly lists files from the 'data' folder
    current_files = set(os.listdir(DATA_FOLDER))
    known_files = set(doc_mapping.keys())
    new_files = current_files - known_files
    deleted_files = known_files - current_files
    updated_files = set()
    for filename in known_files & current_files:
        file_path = os.path.join(DATA_FOLDER, filename)
        if os.path.getmtime(file_path) > doc_mapping[filename]['mtime']:
            updated_files.add(filename)
    return new_files, updated_files, deleted_files


def run_indexing_process():
    global index, documents_list, doc_mapping

    new_files, updated_files, deleted_files = get_file_status()
    if not new_files and not updated_files and not deleted_files:
        print("🔄 No changes detected. Index is up-to-date.")
        if os.path.exists(INDEX_REQUEST_FILE):
            os.remove(INDEX_REQUEST_FILE)
        return

    if deleted_files:
        print(f"🗑️ Deleting {len(deleted_files)} documents from the index.")
        for filename in deleted_files:
            del doc_mapping[filename]

    files_to_index = new_files | updated_files
    if files_to_index:
        print(f"Indexing {len(files_to_index)} new or updated documents...")
        new_docs_content = []
        new_docs_mapping = {}
        for filename in files_to_index:
            file_path = os.path.join(DATA_FOLDER, filename)
            ext = filename.lower().split(".")[-1]
            try:
                if ext == "txt":
                    content = load_txt(file_path)
                elif ext == "pdf":
                    content = load_pdf(file_path)
                elif ext == "docx":
                    content = load_docx(file_path)
                else:
                    continue
                if content:
                    new_docs_content.append(content)
                    new_docs_mapping[filename] = {'content': content, 'mtime': os.path.getmtime(file_path)}
            except Exception as e:
                print(f"❌ Error processing {filename} due to error: {e}. Skipping.")

        if new_docs_content:
            doc_mapping.update(new_docs_mapping)

    print("Rebuilding FAISS index...")
    index = faiss.IndexFlatL2(embedder.get_sentence_embedding_dimension())
    all_docs_content = [item['content'] for item in doc_mapping.values()]
    if all_docs_content:
        all_embeddings = embedder.encode(all_docs_content, convert_to_numpy=True)
        index.add(all_embeddings)

    faiss.write_index(index, INDEX_FILE)
    with open(MAPPING_FILE, "wb") as f:
        pickle.dump(doc_mapping, f)
    documents_list = all_docs_content
    print(f"✅ Indexed {len(doc_mapping)} total documents.")
    current_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    with open(TIMESTAMP_FILE, "w") as f:
        f.write(current_time)
    if os.path.exists(INDEX_REQUEST_FILE):
        os.remove(INDEX_REQUEST_FILE)


def worker_thread():
    while True:
        if os.path.exists(INDEX_REQUEST_FILE):
            print("Indexing request received. Starting process...")
            run_indexing_process()
            load_index_and_docs()
        time.sleep(5)


def init_excel():
    if not os.path.exists(EXCEL_LOG):
        wb = Workbook()
        ws = wb.active
        ws.title = "ChatLogs"
        ws.append(["Session ID", "Date", "Time", "IP Address", "User Message", "Bot Response"])
        wb.save(EXCEL_LOG)


def log_to_excel(session_id, ip, user_msg, bot_resp):
    try:
        wb = load_workbook(EXCEL_LOG)
        ws = wb["ChatLogs"]
        now = datetime.datetime.now()
        date_str = now.strftime("%Y-%m-%d")
        time_str = now.strftime("%H:%M:%S")
        ws.append([session_id, date_str, time_str, ip, user_msg, bot_resp])
        wb.save(EXCEL_LOG)
    except Exception as e:
        print(f"Failed to log to Excel: {e}")


def search_docs(query, top_k=3):
    if not index: return "No documents are indexed. Please run the indexing script."
    query_vec = embedder.encode([query])
    D, I = index.search(query_vec, top_k)
    results = [documents_list[i] for i in I[0] if i < len(documents_list)]
    return "\n".join(results)


@app.before_request
def make_session_permanent():
    session.permanent = True
    if "session_id" not in session:
        session["session_id"] = str(uuid.uuid4())
    if "chat_history" not in session:
        session["chat_history"] = []


@app.route("/")
def home():
    return render_template("index.html", chat_history=session.get("chat_history", []))


@app.route("/chat", methods=["POST"])
def chat():
    user_message = request.json.get("message")
    if not user_message: return jsonify({"error": "No message provided"}), 400

    # Simple check for conversational greetings and other phrases
    greetings = ["hi", "hello", "hey", "مرحبا"]
    how_are_you_phrases = ["how are you doing", "how are you", "كيف حالك"]
    role_phrases = ["what is your role", "what is your purpose", "ما هي وظيفتك"]

    if user_message.lower() in greetings:
        bot_reply = "Hello there! How can I help you today?"
        if user_message.lower() == "مرحبا":
            bot_reply = "أهلاً بك! كيف يمكنني مساعدتك اليوم؟"
    elif user_message.lower() in how_are_you_phrases:
        bot_reply = "I'm doing great, thank you for asking! I'm ready to assist you."
        if user_message.lower() == "كيف حالك":
            bot_reply = "أنا بخير، شكراً لسؤالك! أنا جاهز للمساعدة."
    elif user_message.lower() in role_phrases:
        bot_reply = "I am an AI assistant designed to answer your questions based on the documents you have uploaded."
        if user_message.lower() == "ما هي وظيفتك":
            bot_reply = "أنا مساعد ذكي مصمم للإجابة على أسئلتك بناءً على المستندات التي قمت بتحميلها."
    else:
        # The RAG logic for searching and generating a response
        context = search_docs(user_message)
        if "No documents" in context:
            bot_reply = context
        else:
            relevance_prompt = f"Given the following context, can you answer the question? Answer 'Yes' or 'No'.\n\nContext:\n{context}\n\nQuestion: {user_message}\n\nAnswer:"
            relevance_check = openai.chat.completions.create(
                model="gpt-4o-mini",
                messages=[
                    {"role": "system", "content": "You are a helpful assistant that only answers 'Yes' or 'No'."},
                    {"role": "user", "content": relevance_prompt}
                ]
            )
            is_relevant = relevance_check.choices[0].message.content.strip().lower()

            if is_relevant == "no":
                bot_reply = "I'm sorry, I cannot answer that question. It appears to be outside of my trained knowledge."
            else:
                prompt = f"Context:\n{context}\n\nQuestion: {user_message}\nAnswer:"
                completion = openai.chat.completions.create(
                    model="gpt-4o-mini",
                    messages=[
                        {"role": "system", "content": system_prompt},
                        {"role": "user", "content": prompt}
                    ]
                )
                bot_reply = completion.choices[0].message.content.strip()

    session["chat_history"].append({"role": "user", "text": user_message})
    session["chat_history"].append({"role": "bot", "text": bot_reply})
    session.modified = True
    user_ip = request.remote_addr or "unknown"
    log_to_excel(session["session_id"], user_ip, user_message, bot_reply)
    return jsonify({"reply": bot_reply})


@app.route("/reset", methods=["POST"])
def reset_chat():
    session["chat_history"] = []
    session.modified = True
    return jsonify({"status": "reset"})


@app.route('/training')
def training_dashboard():
    files = []
    if os.path.exists(DATA_FOLDER):
        for filename in os.listdir(DATA_FOLDER):
            file_path = os.path.join(DATA_FOLDER, filename)
            if os.path.isfile(file_path):
                file_stats = os.stat(file_path)
                files.append({
                    'name': filename, 'size': f"{file_stats.st_size / 1024:.2f} KB",
                    'extension': os.path.splitext(filename)[1],
                    'date': datetime.datetime.fromtimestamp(file_stats.st_mtime).strftime('%Y-%m-%d %H:%M:%S')
                })
    last_trained = get_last_training_timestamp()
    return render_template('training.html', files=files, last_trained=last_trained)


def get_last_training_timestamp():
    if os.path.exists(TIMESTAMP_FILE):
        with open(TIMESTAMP_FILE, "r") as f:
            return f.read().strip()
    return "N/A"


@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files: return jsonify({'error': 'No file part'}), 400
    file = request.files['file']
    if file.filename == '': return jsonify({'error': 'No selected file'}), 400
    if file:
        file.save(os.path.join(DATA_FOLDER, file.filename))
        return redirect(url_for('training_dashboard'))


@app.route('/delete/<filename>')
def delete_file(filename):
    file_path = os.path.join(DATA_FOLDER, filename)
    if os.path.exists(file_path): os.remove(file_path)
    return redirect(url_for('training_dashboard'))


@app.route('/delete_all')
def delete_all_files():
    for filename in os.listdir(DATA_FOLDER):
        file_path = os.path.join(DATA_FOLDER, filename)
        if os.path.isfile(file_path): os.remove(file_path)
    if os.path.exists(INDEX_FILE): os.remove(INDEX_FILE)
    if os.path.exists(MAPPING_FILE): os.remove(MAPPING_FILE)
    if os.path.exists(TIMESTAMP_FILE): os.remove(TIMESTAMP_FILE)
    if os.path.exists(INDEX_REQUEST_FILE): os.remove(INDEX_REQUEST_FILE)
    return redirect(url_for('training_dashboard'))


@app.route('/train')
def train_indexing():
    try:
        if os.path.exists(INDEX_REQUEST_FILE):
            return jsonify({'error': 'An indexing job is already in progress. Please wait.'}), 409

        with open(INDEX_REQUEST_FILE, "w") as f:
            f.write("start")
        return jsonify({'message': 'Indexing process started.'})
    except Exception as e:
        return jsonify({'error': f'Failed to start indexing: {e}'}), 500


@app.route('/check_status')
def check_status():
    if os.path.exists(INDEX_REQUEST_FILE):
        return jsonify({'status': 'in_progress'})
    else:
        return jsonify({'status': 'completed', 'last_trained': get_last_training_timestamp()})


# This is the code that is executed when Gunicorn starts the application.
if not os.path.exists(DATA_FOLDER):
    os.makedirs(DATA_FOLDER)
load_index_and_docs()
load_system_prompt()
threading.Thread(target=worker_thread, daemon=True).start()

# This block is only for local development. Gunicorn will not execute it.
if __name__ == '__main__':
    app.run(debug=True)
